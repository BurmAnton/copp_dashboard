import math
from openpyxl import load_workbook
from datetime import datetime

from datacenter.models import Citizen, DisabilityType

def get_sheet(form):
    workbook = load_workbook(form.cleaned_data['import_file'])
    sheet = workbook.active
    return sheet

def cheak_col_match(sheet, fields_names_set):
    i = 0
    col_count = sheet.max_column
    sheet_fields = []
    sheet_col = {}
    if sheet[f"A2"].value is None:
        return ['Import', 'EmptySheet']
    try:
        for col_header in range(1, col_count+1):
            if sheet.cell(row=1,column=col_header).value is not None:
                sheet_fields.append(sheet.cell(row=1,column=col_header).value)
                sheet_col[col_header] = sheet.cell(row=1,column=col_header).value
        missing_fields = []
        for field in fields_names_set:
            if field not in sheet_fields:
                missing_fields.append(field)
        if len(missing_fields) != 0:
            return ['Import', 'MissingFieldsError', missing_fields]
    except IndexError:
            return ['Import', 'IndexError']
    return [True, sheet_col]

def load_worksheet_dict(sheet, fields_names_set):
    row_count = sheet.max_row
    sheet_dict = {}
    for col in fields_names_set:
        sheet_dict[fields_names_set[col]] = []
        for row in range(2, row_count+1): 
            snils = sheet[f"A{row}"].value
            if snils != None:
                cell_value = sheet.cell(row=row,column=col).value
                try: cell_value = str(math.floor(cell_value))
                except (ValueError, TypeError): pass
                sheet_dict[fields_names_set[col]].append(cell_value)
    return sheet_dict

def participants(form, event):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return ['Import', 'IndexError']

    #Требуемые поля таблицы
    fields_names = {
        'Фамилия', 'Имя', 'Отчество',
        'Пол', 'Дата рождения', 'Email',
        'Телефон', 'СНИЛС', 'Гражданин РФ',
        'Образование', 'Занятость в момент регистрации', 'ОВЗ'
    }

    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names[0] != True:
        return cheak_col_names

    sheet_dict = load_worksheet_dict(sheet, cheak_col_names[1])
    
    missing_fields = []
    new_citizen_count = 0
    added_participants_count = 0
    for row in range(len(sheet_dict['СНИЛС'])):
        participant = load_participant(sheet_dict, row)
        if participant[0] == 'OK':
            if participant[1]: new_citizen_count += 1
            else: added_participants_count += 1
            event.participants.add(participant[2])
            event.save()
        elif participant[0] == 'MissingField':
            missing_fields.append(participant)
    return ['Import', 'OK', new_citizen_count, added_participants_count, missing_fields]

def load_participant(sheet, row):
    missing_fields = []
    snils_number = sheet["СНИЛС"][row]
    if snils_number != None: snils_number = snils_number.replace(" ", "")
    last_name = sheet["Фамилия"][row].replace(" ", "")
    if last_name == "": missing_fields.append("Фамилия")
    first_name = sheet["Имя"][row].replace(" ", "")
    if first_name == "": missing_fields.append("Имя")
    middle_name = sheet["Отчество"][row]
    if middle_name == "": middle_name.replace(" ", "")
    sex = sheet["Пол"][row].replace(" ", "")
    if sex == "Мужской": sex = 'M'
    elif sex == "Женский": sex = 'F'
    else: missing_fields.append("Пол")
    if isinstance(sheet["Дата рождения"][row], datetime):
        birthday = sheet["Дата рождения"][row]
    else: birthday = None
    email = sheet["Email"][row].replace(" ", "")
    if email == "": missing_fields.append("Email")
    phone_number = sheet["Телефон"][row]
    if phone_number != None:
        phone_number = phone_number.replace(" ", "")
    education_type = None
    if sheet["Образование"][row] is not None:
        for education_choice in Citizen.EDUCATION_CHOICES:
            if sheet["Образование"][row] in education_choice[1]: 
                education_type = education_choice[0]
                break
    is_russian_citizen = sheet["Гражданин РФ"][row]
    if is_russian_citizen != None: 
        is_russian_citizen = is_russian_citizen.replace(" ", "")
    is_employed = sheet["Занятость в момент регистрации"][row]
    if is_employed != None: 
        is_employed = is_employed.replace(" ", "")
    disability_types = None
    if sheet["ОВЗ"][row] != None:
        disability_types = sheet["ОВЗ"][row].replace("  ", " ").strip().split(", ")
        disability_types = DisabilityType.objects.filter(name__in=disability_types)

    if len(missing_fields) == 0:
        citizen, is_new = Citizen.objects.get_or_create(email=email)
        citizen.snils_number = snils_number
        citizen.last_name = last_name
        citizen.first_name = first_name
        citizen.middle_name = middle_name
        citizen.birthday = birthday
        citizen.sex = sex
        citizen.email = email
        citizen.phone_number = phone_number
        citizen.education_type = education_type
        if is_russian_citizen != "Да": citizen.is_russian_citizen = False
        if is_employed == "Да": citizen.is_employed = True
        if disability_types != None: 
            citizen.disability_type.add(*disability_types)
        citizen.save()
        return ['OK', is_new, citizen]
    return ['MissingField', missing_fields, email]
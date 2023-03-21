from calendar import monthrange
import json
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from django.shortcuts import get_object_or_404, render
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.http import HttpResponseRedirect
from django.utils.encoding import escape_uri_path

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
import pandas as pd

from datacenter.models import  Citizen, EducationProgram, Event, EventType, \
                               Group, Project, DisabilityType, Competence
from reports.models import Tag, ReportTemplate, ReportField, TimeInterval

# Create your views here.
@csrf_exempt
def generate_report(request, report_id):
    if request.method == 'POST':
        report = get_object_or_404(ReportTemplate, id=report_id)
        wb = Workbook()
        ws = wb.active
        ws.title = report.name
        row = 2
        for field in report.fields.all():
            column = 2
            ws.cell(row=row, column=column-1, value=field.name)
            for interval in report.intervals.all():
                end_date = request.POST[f'end_date_{interval.id}']
                end_date = get_end_date(end_date, interval)
                ws.cell(row=1, column=column, value=f'{interval} ({end_date.strftime("%d/%m/%Y")})')
                if interval.interval_type != 'GRWNGA':
                    if interval.interval_type == 'GRWNG':
                        start_date = request.POST[f'start_date_{interval.id}']
                        start_date = get_start_date(start_date, interval)
                    elif interval.interval_type == 'SNGL':
                        start_date = end_date
                        start_date = start_date.replace(day=1)
                        if interval.period == 'QRTR':
                            start_date = start_date.replace(
                                month=start_date.month - 2
                            )
                        elif interval.period == 'YEAR':
                            start_date = start_date.replace(month=1)
                    ws.cell(
                        row=1, 
                        column=column,
                        value=f'{interval} ({start_date.strftime("%d/%m/%Y")}-{end_date.strftime("%d/%m/%Y")})'
                    )
                    cell_value = count_filed_data(field, end_date, start_date)
                else:
                    cell_value = count_filed_data(field, end_date)
                ws.cell(row=row, column=column, value=cell_value)
                column += 1
            row += 1
        wb.template = False
        wb.save(f'{report.name}.xlsx')
        response = HttpResponse(content=save_virtual_workbook(wb), 
                                content_type='application/vnd.openxmlformats-\
                                officedocument.spreadsheetml.sheet'
                    )
        response['Content-Disposition'] = "attachment; filename=" + escape_uri_path(f'{report.name}.xlsx')
        return response

def get_end_date(end_date, interval):
    if interval.period == 'QRTR':
        if 'Q1' in end_date:
            quarter = 1
        elif 'Q2' in end_date:
            quarter = 2
        elif 'Q3' in end_date:
            quarter = 3
        elif 'Q4' in end_date:
            quarter = 4
        year = end_date.split(" ")[1]
        last_month_of_quarter = 3 * quarter
        end_date = date(
                        year, 
                        last_month_of_quarter, 
                        monthrange(year, last_month_of_quarter)[1]
                    )
    elif interval.period == 'DTS':
        end_date = pd.to_datetime(end_date)
    else:
        end_date = pd.to_datetime(end_date)
        end_date = end_date.replace(day=monthrange(end_date.year, end_date.month)[1])
        if interval.period == 'YEAR':
            end_date = end_date.replace(month=12)
    return end_date

def get_start_date(start_date, interval):
    if interval.period == 'QRTR':
        if 'Q1' in start_date:
            quarter = 1
        elif 'Q2' in start_date:
            quarter = 2
        elif 'Q3' in start_date:
            quarter = 3
        elif 'Q4' in start_date:
            quarter = 4
        year = start_date.split(" ")[1]
        first_month_of_quarter = 3 * quarter - 2
        start_date = date(year,first_month_of_quarter, 1)
    else:
        start_date = pd.to_datetime(start_date)
    return start_date

def count_filed_data(field, end_date, start_date=None):
    if field.field_type == 'VNTSPPL' or field.field_type == 'PRGMPPL':
        citizens = Citizen.objects.all()
        if len(field.disabilities.all()) != 0:
            citizens.filter(disability_type__in=field.disabilities.all())
        if field.sex != 'A': citizens.filter(sex=field.sex)
        if field.age_limit_min != None:
            age_limit_min = date.today()-relativedelta(years=age_limit_min)
            citizens.filter(birthday__gte=age_limit_min)
        if field.age_limit_max != None:
            age_limit_max = date.today()-relativedelta(years=age_limit_max)
            citizens.filter(birthday__lte=age_limit_max)
    if field.field_type == 'PRGMPPL':
        groups = Group.objects.all()
        if len(field.projects.all()) != 0:
            groups.filter(project__in=field.projects.all())
        if len(field.competencies.all()) != 0:
            groups.filter(
                education_program__competence__in=field.competencies.all()
            )
        if len(field.tags.all()) != 0:
            groups.filter(tags__in=field.tags.all())
        if len(field.stop_tags.all()) != 0:
            groups.exclude(tags__in=field.stop_tags.all())
        groups = groups.filter(end_date__lte=end_date)
        if start_date != None:
            groups = groups.filter(start_date__gte=start_date)
        data = citizens.filter(groups__in=groups)
    elif field.field_type == 'PRGM':
        programs = EducationProgram.objects.all()
        if len(field.competencies.all()) != 0:
            programs.filter(
                competence__in=field.competencies.all()
            )
        if len(field.tags.all()) != 0:
            programs.filter(tags__in=field.tags.all())
        if len(field.stop_tags.all()) != 0:
            programs.exclude(tags__in=field.stop_tags.all())
        programs = programs.filter(creation_date__lte=end_date)
        if start_date != None:
            programs = programs.filter(creation_date__gte=start_date)
        data = programs
    elif field.field_type == 'VNTSPPL' or field.field_type == 'VNTS':
        events = Event.objects.all()
        if len(field.projects.all()) != 0:
            events.filter(project__in=field.projects.all())
        if len(field.event_types.all()) != 0:
            events.filter(event_type__in=field.event_types.all())
        if len(field.tags.all()) != 0:
            events.filter(tags__in=field.tags.all())
        if len(field.stop_tags.all()) != 0:
            events.exclude(tags__in=field.stop_tags.all())
        events = events.filter(end_date__lte=end_date)
        if start_date != None:
            events = events.filter(start_date__gte=start_date)
        if field.field_type == 'VNTSPPL':
            data = citizens.filter(events__in=events)
        elif field.field_type == 'VNTS':
            data = events
    return data.count()

@csrf_exempt
def reports_page(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        name = data.get("name", "")
        report = ReportTemplate(name=name)
        report.save()
        fields_count = int(data.get("fields_count", ""))
        intervals_count = int(data.get("periods_count", ""))
        for interval_number in range(intervals_count):
            interval = TimeInterval(
                report=report,
                period=data.get(f'period_type_{interval_number + 1}'),
                interval_type=data.get(f'interval_type_{interval_number + 1}')
            )
            interval.save()
        for field_number in range(fields_count):
            sex=data.get(f'sex_{field_number + 1}')
            if sex[0] == "Любой": sex = 'A'
            elif sex[0] == "Мужской": sex = 'M'
            elif sex[0] == "Женский": sex = 'F'
            field = ReportField(
                item_number=field_number,
                name=data.get(f'field_name_{field_number + 1}'),
                field_type=data.get(f'field_type_{field_number + 1}'),
                sex=sex,
                report=report,
            )
            age_limit_min = data.get(f'age_limit_min_{field_number+1}')
            age_limit_max = data.get(f'age_limit_max_{field_number+1}')
            if age_limit_min is not None:
                field.age_limit_min=int(age_limit_min)
            if age_limit_max is not None:
                field.age_limit_max=int(age_limit_max)
            field.save()
            field.tags.set(Tag.objects.filter(
                    name__in=data.get(f'tags_{field_number + 1}')
                ))
            field.stop_tags.set(Tag.objects.filter(
                name__in=data.get(f'stop_tags_{field_number + 1}')
            ))
            field.competencies.set(Competence.objects.filter(
                name__in=data.get(f'competencies_{field_number + 1}')
            ))
            field.disabilities.set(DisabilityType.objects.filter(
                name__in=data.get(f'disabilities_{field_number + 1}')
            ))
            field.projects.set(Project.objects.filter(
                name__in=data.get(f'projects_{field_number + 1}')
            ))
            field.event_types.set(EventType.objects.filter(
                name__in=data.get(f'event_types_{field_number + 1}')
            ))
            
        return JsonResponse({"message": "OK"}, status=201)
    reports = ReportTemplate.objects.all()
    message = None
    disabilities = DisabilityType.objects.all()
    competencies = Competence.objects.all()
    event_types = EventType.objects.all()
    projects = Project.objects.all()
    tags = Tag.objects.all()

    period_types = TimeInterval.PERIOD_TYPES
    interval_types = TimeInterval.INTERVAL_TYPES
    field_types = ReportField.FIELD_TYPES

    return render(request, "reports/reports_page.html", {
        "reports": reports,
        "tags": tags, 
        "projects": projects, 
        "competencies": competencies, 
        "event_types": event_types, 
        "disabilities": disabilities,
        "period_types": period_types,
        "interval_types": interval_types,
        "field_types": field_types,
        "message": message,
        "page_name": "Отчёты | ЦОПП СО"
    })